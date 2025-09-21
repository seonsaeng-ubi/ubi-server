from .models import BigSubject, SmallSubject, Region, Problem, TestSet, RealRegion, StudyRoom
from .helpers import generate_unique_room_no, get_random_practice_problems, get_mockup_problems
from rest_framework.generics import RetrieveUpdateAPIView, ListAPIView, RetrieveAPIView
from .serializers import BigSubjectSerializer, RegionSerializer, ProblemListSerializer, \
    ProblemUpdateSerializer, TestSetSerializer, StudyRoomSerializer, StudyRoomListSerializer
from rest_framework.permissions import IsAuthenticated, AllowAny
from openpyxl_image_loader import SheetImageLoader
from ..utils import StandardResultsSetPagination
from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from rest_framework.status import HTTP_200_OK
from rest_framework.response import Response
from rest_framework.views import APIView
from api.stats.models import Setting
from django.core.files import File
from .forms import ExcelUploadForm
from django.db.models import Value
from django.db.models import Q
import openpyxl
import random
import boto3
import os


class MyStudyRoomListAPIView(ListAPIView):
    serializer_class = StudyRoomListSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return StudyRoom.objects.filter(users=self.request.user).prefetch_related('problems', 'users').order_by('-id')


class AllStudyRoomListAPIView(ListAPIView):
    serializer_class = StudyRoomSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return StudyRoom.objects.prefetch_related('problems', 'users').order_by('-id')


class StudyRoomDetailAPIView(RetrieveAPIView):
    serializer_class = StudyRoomSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        pk = self.kwargs.get('pk')
        study_room = get_object_or_404(StudyRoom.objects.prefetch_related('problems', 'users'), pk=pk)

        # 현재 사용자를 스터디룸에 추가 (이미 있으면 중복 추가 안됨)
        study_room.users.add(self.request.user)

        return study_room


class StudyRoomSearchAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        room_no = request.query_params.get('room_no', '')
        room_no = room_no.strip()
        if room_no == '':
            return Response([], status=200)

        # 접두어로 시작하는 스터디룸 전체 검색
        study_rooms = StudyRoom.objects.filter(room_no__startswith=room_no).order_by('-id')

        # 리스트로 반환 (없으면 빈 리스트)
        serializer = StudyRoomListSerializer(study_rooms, many=True)
        return Response(serializer.data, status=200)


class StudyRoomCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        # 현재 사용자
        user = request.user
        # 지역 필터링
        region = Region.objects.get(id=int(request.data.get('region', '1')))
        # 문제 타입 필터링 (구상형, 즉답형, 모의고사 세트)
        type = request.data.get('type', 'A')
        # 스터디룸 번호 생성
        room_no = generate_unique_room_no()

        study_room = StudyRoom.objects.create(
            region=region,
            type=type,
            room_no=room_no,
        )
        study_room.users.add(user)
        # 질문 세트 분기
        if type == 'C':
            problems = get_mockup_problems(region.id)
        else:
            problems = get_random_practice_problems(region.id, type)
        # 질문 넣기
        study_room.problems.set(problems)
        study_room.save()

        # 여기에 data serialize해서 response 할 것.
        serializer = StudyRoomSerializer(study_room)
        return Response(serializer.data, status=200)


class StudyRoomLeaveAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk, *args, **kwargs):
        try:
            study_room = StudyRoom.objects.get(pk=pk)
        except StudyRoom.DoesNotExist:
            return Response({'detail': '해당 스터디룸을 찾을 수 없습니다.'}, status=404)
        user = request.user

        # 사용자가 해당 스터디룸에 참여하고 있는지 확인
        if not study_room.users.filter(id=user.id).exists():
            return Response({'detail': '해당 스터디룸에 참여하고 있지 않습니다.'}, status=400)

        # 사용자를 스터디룸에서 제거
        study_room.users.remove(user)

        # 모든 사용자가 나갔는지 확인
        if study_room.users.count() == 0:
            # 모든 사용자가 나갔으면 스터디룸 삭제
            study_room.delete()
            return Response({'detail': '스터디룸에서 나갔습니다. 모든 참여자가 나가서 스터디룸이 삭제되었습니다.'}, status=200)

        return Response({'detail': '스터디룸에서 나갔습니다.'}, status=200)


class SubjectAPIView(ListAPIView):
    serializer_class = BigSubjectSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return BigSubject.objects.all()


class RegionAPIView(ListAPIView):
    serializer_class = RegionSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Region.objects.all()[:3]


# 스크랩 취소 / 하기
class ScrapUpdateAPIView(RetrieveUpdateAPIView):
    serializer_class = ProblemUpdateSerializer
    permission_classes = [IsAuthenticated]
    queryset = Problem.objects.all()
    allowed_methods = ['put', 'get']
    lookup_field = 'pk'

    def get_object(self):
        return Problem.objects.get(id=self.kwargs['pk'])


# 서버 세팅 확인
def get_setting():
    setting_model = Setting.objects.last()
    is_blocked = setting_model.is_blocked
    return is_blocked


# 연습 문제 / 지역 필터링, 구상 or 즉답형 필터링
class AllPracticeProblemListAPIVIew(ListAPIView):
    serializer_class = ProblemListSerializer
    permission_classes = [IsAuthenticated]
    model = Problem

    def get_queryset(self):
        # 세팅 값 불러오기
        is_blocked = get_setting()
        if is_blocked is True:
            return Problem.objects.none()
        else:
            # 지역 필터링
            regions = Region.objects.all()
            pyeonggawon = regions.get(title='평가원')
            gongtong = regions.get(title='공통')

            region = regions.get(id=int(self.request.query_params.get('region', '1')))

            # 문제 타입 필터링 (구상형, 즉답형, 추가 질문)
            type = self.request.query_params.get('type', 'A')

            # 평가원일 경우
            if region.id == pyeonggawon.id:
                problems = Problem.objects.filter(
                    Q(region=region) & Q(type=type) & Q(problem_type='P')
                )

            # 그 외의 경우
            else:
                problems = Problem.objects.filter(
                    (Q(region=region) | Q(region=gongtong)) & Q(type=type) & Q(problem_type='P')
                )
            return problems.order_by('number')


# 랜덤 연습 문제 / 지역 필터링, 구상 or 즉답형 필터링
class AllRandomPracticeProblemListAPIVIew(ListAPIView):
    serializer_class = ProblemListSerializer
    permission_classes = [IsAuthenticated]
    model = Problem

    def get_queryset(self):
        # 세팅 값 불러오기
        is_blocked = get_setting()
        if is_blocked is True:
            return Problem.objects.none()
        else:
            # 지역 필터링
            regions = Region.objects.all()
            pyeonggawon = regions.get(title='평가원')
            gongtong = regions.get(title='공통')

            region = regions.get(id=int(self.request.query_params.get('region', '1')))

            # 문제 타입 필터링 (구상형, 즉답형, 추가 질문)
            type = self.request.query_params.get('type', 'A')

            # 평가원일 경우
            if region.id == pyeonggawon.id:
                problems = Problem.objects.filter(
                    Q(region=region) & Q(type=type) & Q(problem_type='P')
                )

            # 그 외의 경우
            else:
                problems = Problem.objects.filter(
                    (Q(region=region) | Q(region=gongtong)) & Q(type=type) & Q(problem_type='P')
                )
            return problems.order_by('?')


# 문제 디테일
class ProblemDetailAPIView(RetrieveAPIView):
    serializer_class = ProblemListSerializer
    permission_classes = [IsAuthenticated]
    queryset = Problem.objects.all().prefetch_related('scrapped_users')
    lookup_field = 'pk'

    def get_object(self):
        return Problem.objects.get(id=self.kwargs['pk'])


# 새로운 기출문제 리스트 (연도별)
class NewRealProblemListAPIView(APIView):
    def get(self, request, *args, **kwargs):
        # 세팅 값 불러오기
        is_blocked = get_setting()
        if is_blocked is True:
            return Problem.objects.none()
        else:
            # 지역 필터링 및 기출 출제 지역 가져오기
            regions = Region.objects.all()
            region = regions.get(id=int(self.request.query_params.get('region', '1')))
            real_region = RealRegion.objects.get(title=region.title)
            # 기출 출제 지역 바탕 필터링
            problems = Problem.objects.filter(
                Q(real_region=real_region) & Q(problem_type='A')
            ).order_by('number')

            years = list(set(problems.values_list('year', flat=True)))
            years.reverse()
            return_list = []

            for year in years:
                temp_problems = problems.filter(year=year)
                serialized_temp_problems = ProblemListSerializer(data=temp_problems, many=True,
                                                                 context={'request': request})
                if serialized_temp_problems.is_valid():
                    pass
                serialized_problems = serialized_temp_problems.data
                return_list.append({'year': year, 'problems': serialized_problems})
            return Response(return_list, status=HTTP_200_OK)


# 여기부터 오답노트 리스트
class ScrappedProblemListAPIView(ListAPIView):
    serializer_class = ProblemListSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    model = Problem

    def get_queryset(self):
        # 세팅 값 불러오기
        is_blocked = get_setting()
        if is_blocked is True:
            return Problem.objects.none()
        else:
            big_subject_id = int(self.request.query_params.get('bigSubject', '0'))
            small_subject_id = int(self.request.query_params.get('smallSubject', '0'))
            problems = Problem.objects.order_by('number').prefetch_related('scrapped_users', 'big_subject', 'small_subject')
            # 대주제 전체보기일 경우
            if big_subject_id == 0:
                # ㄹㅇ 전체보기일 경우
                if small_subject_id == 0:
                    problem_list = problems.filter(
                        scrapped_users__in=[self.request.user.id]
                    )
                    return problem_list
                # 소주제만 선택할 경우
                else:
                    problem_list = problems.filter(
                        Q(scrapped_users__in=[self.request.user.id]) & Q(small_subject__in=[small_subject_id])
                    )
                    return problem_list
            else:
                if small_subject_id == 0:
                    problem_list = problems.filter(
                        Q(scrapped_users__in=[self.request.user.id]) &
                        Q(big_subject_id=big_subject_id)
                    )
                    return problem_list
                else:
                    problem_list = problems.filter(
                        Q(scrapped_users__in=[self.request.user.id]) &
                        Q(big_subject_id=big_subject_id) &
                        Q(small_subject__in=[small_subject_id])
                    )
                    return problem_list


# 마이페이지 검색
class SearchProblem(ListAPIView):
    serializer_class = ProblemListSerializer
    permission_classes = [IsAuthenticated]
    model = Problem

    def get_queryset(self):
        # 세팅 값 불러오기
        is_blocked = get_setting()
        if is_blocked is True:
            return Problem.objects.none()
        else:
            number = self.request.query_params.get('number', '  ')
            return Problem.objects.filter(number__startswith=number).order_by('number')


# 실전 모의고사 설명
class TestSetView(RetrieveAPIView):
    serializer_class = TestSetSerializer
    permission_classes = [AllowAny]
    model = TestSet
    lookup_field = 'pk'
    queryset = TestSet.objects.all()

    def get_object(self):
        region = self.kwargs['pk']
        return TestSet.objects.get(region_id=region)


# 실전 모의고사 리스트
class TestSetListView(ListAPIView):
    serializer_class = ProblemListSerializer
    permission_classes = [IsAuthenticated]
    model = Problem

    def get_queryset(self):
        # 세팅 값 불러오기
        is_blocked = get_setting()
        if is_blocked is True:
            return Problem.objects.none()
        else:
            # 지역 필터링
            regions = Region.objects.all()
            seoul = regions.get(title='서울')
            gyeonggi = regions.get(title='경기')
            # sejong = regions.get(title='세종')
            pyeonggawon = regions.get(title='평가원')
            gongtong = regions.get(title='공통')
            region = regions.get(id=int(self.request.query_params.get('region', '1')))

            # 공통이거나 특정 지역일 경우 우선 필터링
            if region.id == pyeonggawon.id:
                problems = Problem.objects.filter(region=region)
            else:
                problems = Problem.objects.filter(Q(region=region) | Q(region=gongtong))
            # 서울일 경우
            if region.id == seoul.id:
                # 구상형 3문제, 즉답형 3, 추가 질문 2
                conception = problems.filter(type='A').order_by('?')[:2]
                immediate = problems.filter(type='B').order_by('?')[:3]
                additional = problems.filter(type='C').order_by('?')[:2]

                type_a = conception.annotate(custom_order=Value(1))
                type_b = immediate.annotate(custom_order=Value(2))
                type_c = additional.annotate(custom_order=Value(3))
                temp_result = type_a.union(type_b).order_by('custom_order')
                result = temp_result.union(type_c).order_by('custom_order')

                return result
            # 경기일 경우
            elif region.id == gyeonggi.id:
                # 구상형 3문제, 즉답형 2
                conception = problems.filter(type='A').order_by('?')[:3]
                immediate = problems.filter(type='B').order_by('?')[:2]

                type_a = conception.annotate(custom_order=Value(1))
                type_b = immediate.annotate(custom_order=Value(2))
                result = type_a.union(type_b).order_by('custom_order')

                return result
            # 평가원일 경우
            else:
                # 구상형 3문제, 즉답형 1
                conception = problems.filter(type='A').order_by('?')[:3]
                immediate = problems.filter(type='B').order_by('?')[:1]

                type_a = conception.annotate(custom_order=Value(1))
                type_b = immediate.annotate(custom_order=Value(2))
                result = type_a.union(type_b).order_by('custom_order')

                return result


# 실제 문제가 있는 지역에 관한 API
class ActiveRegionAPIView(APIView):
    def get(self, request, *args, **kwargs):
        # 문제 다 가져오기
        problems = Problem.objects.values('region')
        problem_regions = list(problems)

        return Response(problem_regions, status=HTTP_200_OK)


# 엑셀 업로드
def upload_excel(request):
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_files = request.FILES.getlist('excel_file')
            for excel_file in excel_files:
                wb = openpyxl.load_workbook(excel_file)
                # 워크시트 첫 장 불러오기
                ws = wb.worksheets[0]

                number = ws['B5'].value
                writer = ws['C5'].value
                type = ws['D5'].value
                region = ws['E5'].value
                is_practice = ws['F5'].value
                big_subject = ws['G5'].value
                small_subject = ws['H5'].value
                has_picture = ws['J5'].value
                real_region = ws['K5'].value
                title = ws['L5'].value
                year = ws['M5'].value
                presentation = ws['G8'].value
                question = ws['K8'].value
                answer = ws['L8'].value

                if Problem.objects.filter(number=int(number)).exists():
                    pass
                else:
                    big_subject_obj, big_subject_created = BigSubject.objects.get_or_create(title=big_subject)
                    small_subject_obj, small_subject_created = SmallSubject.objects.get_or_create(title=small_subject,
                                                                                                  big_subject=big_subject_obj)
                    if small_subject_created:
                        chars = '0123456789ABCDEF'
                        small_subject_obj.color = ''.join(random.choice(chars) for i in range(6))
                        small_subject_obj.save()

                    new_problem = Problem.objects.create(
                        number=int(number),
                        writer=writer,
                        type='A' if type == '구상형' else 'B',
                        region=Region.objects.get(title=region),
                        real_region=RealRegion.objects.get(title=real_region) if real_region is not None else None,
                        problem_type='P' if is_practice == '예상문제' else 'A',
                        title=title,
                        year=int(year) if year is not None else None,
                        presentation=presentation,
                        question=question,
                        answer=answer,
                        has_image=False if has_picture == '없음' else True,
                        big_subject=big_subject_obj,
                    )
                    new_problem.small_subject.add(small_subject_obj)

                    img = None
                    if has_picture == '없음':
                        new_problem.save()
                    else:
                        # 이미지 로더
                        image_loader = SheetImageLoader(ws)

                        alphabets = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

                        for image in ws._images:
                            col = image.anchor._from.col
                            row = image.anchor._from.row + 1

                            img = image_loader.get(alphabets[col] + str(row))
                            img.save(str(number) + '.jpeg')

                        with open(str(number) + '.png', 'rb') as f:
                            image_file = File(f)
                            s3 = boto3.client(
                                's3', aws_access_key_id=os.getenv('S3_ACCESS_KEY_ID'),
                                aws_secret_access_key=os.getenv('S3_SECRET_ACCESS_KEY'),
                                region_name='ap-northeast-2'
                            )
                            s3.upload_fileobj(image_file, 'ubi-s3-bucket', str(number) + '.jpeg')

                            new_problem.image_url = 'https://ubi-s3-bucket.s3.ap-northeast-2.amazonaws.com/' + str(
                                number) + '.jpeg'

                        new_problem.save()

            return redirect('success_page')  # 성공시 리다이렉트할 URL 설정

    else:
        form = ExcelUploadForm()
    return render(request, 'upload_excel.html', {'form': form})


def success_view(request):
    return render(request, 'success.html')
